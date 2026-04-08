from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from io import BytesIO
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=16)
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
DEFAULT_TANK_CAPACITY_GALLONS = 200.0


ColumnSpec = tuple[str, Callable[[dict], object]]


def build_motive_snapshot_workbook(snapshot: dict) -> bytes:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"

    drivers = snapshot.get("drivers") or []
    vehicles = snapshot.get("vehicles") or []
    datasets = snapshot.get("datasets") or {}
    recent_activity = snapshot.get("recent_activity") or {}
    metrics = snapshot.get("metrics") or {}
    warnings = snapshot.get("warnings") or []
    company = snapshot.get("company") or {}
    windows = snapshot.get("windows") or {}

    driver_name_to_vehicles = _build_driver_name_map(vehicles)
    roster_by_name = {_normalize_name(driver.get("full_name")): driver for driver in drivers if _normalize_name(driver.get("full_name"))}
    fleet_status_rows = sorted(vehicles, key=lambda row: (_fleet_status_rank(row, roster_by_name), str(row.get("number") or "")))

    _write_overview_sheet(
        overview,
        snapshot=snapshot,
        company=company,
        metrics=metrics,
        datasets=datasets,
        warnings=warnings,
        windows=windows,
    )

    _write_table_sheet(
        workbook,
        title="Fleet Status",
        columns=_fleet_status_columns(roster_by_name),
        rows=fleet_status_rows,
        empty_message="No fleet status rows were available for export.",
    )

    _write_table_sheet(
        workbook,
        title="Driver Directory",
        columns=[
            ("Driver ID", lambda row: row.get("id")),
            ("Full Name", lambda row: row.get("full_name")),
            ("Email", lambda row: row.get("email")),
            ("Phone", lambda row: row.get("phone")),
            ("Role", lambda row: row.get("role")),
            ("Status", lambda row: row.get("status")),
            ("Duty Status", lambda row: row.get("duty_status")),
            ("Username", lambda row: row.get("username")),
            ("Time Zone", lambda row: row.get("time_zone")),
            ("ELD Mode", lambda row: row.get("eld_mode")),
            ("Carrier Name", lambda row: row.get("carrier_name")),
            ("Driver Company ID", lambda row: row.get("driver_company_id")),
            ("License Number", lambda row: row.get("license_number")),
            ("License State", lambda row: row.get("license_state")),
            ("Tracked Vehicle Count", lambda row: len(driver_name_to_vehicles.get(_normalize_name(row.get("full_name")), []))),
            ("Tracked Vehicles", lambda row: ", ".join(driver_name_to_vehicles.get(_normalize_name(row.get("full_name")), []))),
        ],
        rows=drivers,
        empty_message="No driver directory rows were returned by Motive.",
    )

    _write_table_sheet(
        workbook,
        title="Driver Tracking",
        columns=_driver_tracking_columns(roster_by_name),
        rows=vehicles,
        empty_message="No vehicle tracking rows were available for export.",
    )

    _write_table_sheet(
        workbook,
        title="Vehicles",
        columns=_vehicle_columns(),
        rows=vehicles,
        empty_message="No vehicle rows were available for export.",
    )

    activity_sheets = [
        ("Recent Faults", _fault_columns(), recent_activity.get("fault_codes") or [], "No recent fault codes in the current snapshot."),
        ("Recent Safety", _performance_columns(), recent_activity.get("performance_events") or [], "No recent safety events in the current snapshot."),
        ("Recent Driving", _driving_columns(), recent_activity.get("driving_periods") or [], "No recent driving periods in the current snapshot."),
        ("Recent Idling", _idle_columns(), recent_activity.get("idle_events") or [], "No recent idle events in the current snapshot."),
        ("Recent IFTA", _ifta_columns(), recent_activity.get("ifta_trips") or [], "No recent IFTA trips in the current snapshot."),
        ("Fuel Purchases", _fuel_purchase_columns(), recent_activity.get("fuel_purchases") or [], "No recent fuel purchases in the current snapshot."),
        ("Inspections", _inspection_columns(), recent_activity.get("inspection_reports") or [], "No recent inspection reports in the current snapshot."),
        ("Forms", _form_columns(), recent_activity.get("form_entries") or [], "No recent form entries in the current snapshot."),
        ("Driver Scores", _driver_score_columns(), recent_activity.get("driver_scores") or [], "No driver score rows in the current snapshot."),
    ]
    for title, columns, rows, empty_message in activity_sheets:
        _write_table_sheet(workbook, title=title, columns=columns, rows=rows, empty_message=empty_message)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.getvalue()


def _write_overview_sheet(
    sheet: Worksheet,
    *,
    snapshot: dict,
    company: dict,
    metrics: dict,
    datasets: dict,
    warnings: list[str],
    windows: dict,
) -> None:
    generated_at = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    sheet["A1"] = "Motive Fleet Tracking Export"
    sheet["A1"].font = TITLE_FONT
    sheet["A3"] = "Generated At"
    sheet["B3"] = generated_at
    sheet["A4"] = "Snapshot Fetched At"
    sheet["B4"] = snapshot.get("fetched_at")
    sheet["A5"] = "Company"
    sheet["B5"] = company.get("name") or metrics.get("company_name")
    sheet["A6"] = "Auth Mode"
    sheet["B6"] = snapshot.get("auth_mode")
    sheet["A7"] = "Configured"
    sheet["B7"] = "Yes" if snapshot.get("configured") else "No"

    for cell in ("A3", "A4", "A5", "A6", "A7"):
        sheet[cell].font = SECTION_FONT

    start_row = 10
    sheet[f"A{start_row}"] = "Export Algorithm"
    sheet[f"A{start_row}"].font = SECTION_FONT
    sheet[f"A{start_row}"].fill = SECTION_FILL
    algorithm_lines = [
        "1. Refresh Tracking to update the shared Motive fleet snapshot cache.",
        "2. Export reads that snapshot and writes one workbook with separate sheets for overview, drivers, vehicles, and recent activity.",
        "3. Driver Tracking rows combine live vehicle telemetry, utilization, safety, IFTA, fuel, and roster data when Motive provides it.",
        "4. If Motive sends only a raw fuel sensor and not a usable fuel percentage, the export keeps that sensor reading visible without inventing a percent.",
        "5. Activity sheets reflect the current Tracking snapshot windows and recent fleet activity lists returned by the backend.",
    ]
    for index, line in enumerate(algorithm_lines, start=start_row + 1):
        sheet[f"A{index}"] = line

    metric_row = start_row + len(algorithm_lines) + 3
    sheet[f"A{metric_row}"] = "Fleet Metrics"
    sheet[f"A{metric_row}"].font = SECTION_FONT
    sheet[f"A{metric_row}"].fill = SECTION_FILL
    sheet.append([])
    sheet.append([])

    current_row = metric_row + 1
    sheet[f"A{current_row}"] = "Metric"
    sheet[f"B{current_row}"] = "Value"
    _style_header_row(sheet, current_row, 2)
    current_row += 1
    for key, value in metrics.items():
        sheet[f"A{current_row}"] = _humanize(key)
        sheet[f"B{current_row}"] = _cell_value(value)
        current_row += 1

    current_row += 1
    sheet[f"A{current_row}"] = "Dataset Availability"
    sheet[f"A{current_row}"].font = SECTION_FONT
    sheet[f"A{current_row}"].fill = SECTION_FILL
    current_row += 1
    sheet[f"A{current_row}"] = "Dataset"
    sheet[f"B{current_row}"] = "Count"
    sheet[f"C{current_row}"] = "Available"
    _style_header_row(sheet, current_row, 3)
    current_row += 1
    for key, value in datasets.items():
        sheet[f"A{current_row}"] = _humanize(key)
        sheet[f"B{current_row}"] = _cell_value((value or {}).get("count"))
        sheet[f"C{current_row}"] = "Yes" if (value or {}).get("available") else "No"
        current_row += 1

    current_row += 1
    sheet[f"A{current_row}"] = "Snapshot Windows"
    sheet[f"A{current_row}"].font = SECTION_FONT
    sheet[f"A{current_row}"].fill = SECTION_FILL
    current_row += 1
    sheet[f"A{current_row}"] = "Window"
    sheet[f"B{current_row}"] = "Value"
    _style_header_row(sheet, current_row, 2)
    current_row += 1
    for key, value in windows.items():
        sheet[f"A{current_row}"] = _humanize(key)
        sheet[f"B{current_row}"] = _cell_value(value)
        current_row += 1

    current_row += 1
    sheet[f"A{current_row}"] = "Warnings"
    sheet[f"A{current_row}"].font = SECTION_FONT
    sheet[f"A{current_row}"].fill = SECTION_FILL
    current_row += 1
    if warnings:
        for warning in warnings:
            sheet[f"A{current_row}"] = warning
            current_row += 1
    else:
        sheet[f"A{current_row}"] = "No warnings were present in the snapshot."

    _apply_sheet_formatting(sheet)


def _write_table_sheet(workbook: Workbook, *, title: str, columns: list[ColumnSpec], rows: list[dict], empty_message: str) -> None:
    sheet = workbook.create_sheet(title=_sheet_title(title))
    headers = [header for header, _ in columns]
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGNMENT

    if rows:
        for row_index, row in enumerate(rows, start=2):
            for column_index, (_, getter) in enumerate(columns, start=1):
                sheet.cell(row=row_index, column=column_index, value=_cell_value(getter(row))).alignment = WRAP_ALIGNMENT
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    else:
        sheet.cell(row=2, column=1, value=empty_message)
        sheet.cell(row=2, column=1).alignment = WRAP_ALIGNMENT

    _apply_sheet_formatting(sheet)


def _fleet_status_columns(roster_by_name: dict[str, dict]) -> list[ColumnSpec]:
    def roster(vehicle: dict) -> dict:
        return roster_by_name.get(_normalize_name(_tracked_driver_name(vehicle)), {})

    return [
        ("Truck #", lambda row: row.get("number")),
        ("Working Now", lambda row: _working_now_label(row, roster(row))),
        ("Movement", lambda row: _movement_label(row)),
        ("Tracked Driver", lambda row: _tracked_driver_name(row)),
        ("Driver Source", lambda row: _tracked_driver_source(row)),
        ("Driver Status", lambda row: roster(row).get("status")),
        ("Duty Status", lambda row: roster(row).get("duty_status")),
        ("Vehicle Status", lambda row: row.get("status")),
        ("Availability", lambda row: row.get("availability_status")),
        ("Fuel %", lambda row: (row.get("location") or {}).get("fuel_level_percent")),
        ("Fuel Gallons (200 Tank)", lambda row: _estimated_fuel_gallons((row.get("location") or {}).get("fuel_level_percent"))),
        ("Fuel State", lambda row: _fuel_state_label((row.get("location") or {}).get("fuel_level_percent"), (row.get("location") or {}).get("fuel_sensor_reading"))),
        ("Fuel Sensor", lambda row: (row.get("location") or {}).get("fuel_sensor_reading")),
        ("Speed MPH", lambda row: (row.get("location") or {}).get("speed_mph")),
        ("Last Update", lambda row: (row.get("location") or {}).get("located_at")),
        ("Age Minutes", lambda row: (row.get("location") or {}).get("age_minutes")),
        ("Active Faults", lambda row: (row.get("fault_summary") or {}).get("active_count")),
        ("Utilization %", lambda row: (row.get("utilization_summary") or {}).get("utilization_percentage")),
        ("Driving Miles 7D", lambda row: (row.get("driving_summary") or {}).get("distance_miles")),
        ("Safety Events 7D", lambda row: (row.get("performance_summary") or {}).get("count")),
        ("IFTA Miles 30D", lambda row: (row.get("ifta_summary") or {}).get("distance_miles")),
        ("City", lambda row: (row.get("location") or {}).get("city")),
        ("State", lambda row: (row.get("location") or {}).get("state")),
        ("Live Address", lambda row: (row.get("location") or {}).get("address")),
        ("VIN", lambda row: row.get("vin")),
        ("Make", lambda row: row.get("make")),
        ("Model", lambda row: row.get("model")),
        ("Year", lambda row: row.get("year")),
        ("Plate", lambda row: row.get("license_plate_number")),
        ("Plate State", lambda row: row.get("license_plate_state")),
    ]


def _driver_tracking_columns(roster_by_name: dict[str, dict]) -> list[ColumnSpec]:
    def roster(vehicle: dict) -> dict:
        return roster_by_name.get(_normalize_name(_tracked_driver_name(vehicle)), {})

    def scorecard(vehicle: dict) -> dict:
        return vehicle.get("driver_scorecard") or {}

    return [
        ("Tracked Driver Name", lambda row: _tracked_driver_name(row)),
        ("Driver Name Source", lambda row: _tracked_driver_source(row)),
        ("Roster Driver ID", lambda row: roster(row).get("id")),
        ("Roster Email", lambda row: roster(row).get("email")),
        ("Roster Phone", lambda row: roster(row).get("phone")),
        ("Roster Role", lambda row: roster(row).get("role")),
        ("Roster Status", lambda row: roster(row).get("status")),
        ("Roster Duty Status", lambda row: roster(row).get("duty_status")),
        ("Vehicle Number", lambda row: row.get("number")),
        ("Vehicle ID", lambda row: row.get("id")),
        ("VIN", lambda row: row.get("vin")),
        ("Make", lambda row: row.get("make")),
        ("Model", lambda row: row.get("model")),
        ("Year", lambda row: row.get("year")),
        ("License Plate", lambda row: row.get("license_plate_number")),
        ("Plate State", lambda row: row.get("license_plate_state")),
        ("Fuel Type", lambda row: row.get("fuel_type")),
        ("Fuel Level %", lambda row: (row.get("location") or {}).get("fuel_level_percent")),
        ("Secondary Fuel %", lambda row: (row.get("location") or {}).get("fuel_secondary_percent")),
        ("Fuel Gallons (200 Tank)", lambda row: _estimated_fuel_gallons((row.get("location") or {}).get("fuel_level_percent"))),
        ("Fuel Sensor Reading", lambda row: (row.get("location") or {}).get("fuel_sensor_reading")),
        ("Range Remaining", lambda row: (row.get("location") or {}).get("range_remaining")),
        ("Status", lambda row: row.get("status")),
        ("Availability Status", lambda row: row.get("availability_status")),
        ("Live Address", lambda row: (row.get("location") or {}).get("address")),
        ("City", lambda row: (row.get("location") or {}).get("city")),
        ("State", lambda row: (row.get("location") or {}).get("state")),
        ("Latitude", lambda row: (row.get("location") or {}).get("lat")),
        ("Longitude", lambda row: (row.get("location") or {}).get("lon")),
        ("Located At", lambda row: (row.get("location") or {}).get("located_at")),
        ("Location Age Minutes", lambda row: (row.get("location") or {}).get("age_minutes")),
        ("Speed MPH", lambda row: (row.get("location") or {}).get("speed_mph")),
        ("Odometer", lambda row: (row.get("location") or {}).get("true_odometer") or (row.get("location") or {}).get("odometer")),
        ("Engine Hours", lambda row: (row.get("location") or {}).get("true_engine_hours") or (row.get("location") or {}).get("engine_hours")),
        ("Moving Now", lambda row: "Yes" if row.get("is_moving") else "No"),
        ("Stale", lambda row: "Yes" if row.get("is_stale") else "No"),
        ("Active Faults", lambda row: (row.get("fault_summary") or {}).get("active_count")),
        ("Total Faults", lambda row: (row.get("fault_summary") or {}).get("count")),
        ("Utilization %", lambda row: (row.get("utilization_summary") or {}).get("utilization_percentage")),
        ("Idle Hours 7D", lambda row: _hours((row.get("idle_summary") or {}).get("duration_seconds"))),
        ("Idle Events 7D", lambda row: (row.get("idle_summary") or {}).get("count")),
        ("Driving Hours 7D", lambda row: _hours((row.get("driving_summary") or {}).get("duration_seconds"))),
        ("Driving Miles 7D", lambda row: (row.get("driving_summary") or {}).get("distance_miles")),
        ("Safety Events 7D", lambda row: (row.get("performance_summary") or {}).get("count")),
        ("Pending Review Events", lambda row: (row.get("performance_summary") or {}).get("pending_review_count")),
        ("Recent Behaviors", lambda row: ", ".join((row.get("performance_summary") or {}).get("behaviors") or [])),
        ("IFTA Trips 30D", lambda row: (row.get("ifta_summary") or {}).get("count")),
        ("IFTA Miles 30D", lambda row: (row.get("ifta_summary") or {}).get("distance_miles")),
        ("Fuel Purchases 30D", lambda row: (row.get("fuel_purchase_summary") or {}).get("count")),
        ("Fuel Purchase Volume", lambda row: (row.get("fuel_purchase_summary") or {}).get("volume_total")),
        ("Fuel Purchase Amount", lambda row: (row.get("fuel_purchase_summary") or {}).get("amount_total")),
        ("Last Fuel Vendor", lambda row: (row.get("fuel_purchase_summary") or {}).get("last_vendor")),
        ("Inspections 30D", lambda row: (row.get("inspection_summary") or {}).get("count")),
        ("Unsafe Inspections", lambda row: (row.get("inspection_summary") or {}).get("unsafe_count")),
        ("Form Entries 30D", lambda row: (row.get("form_summary") or {}).get("count")),
        ("Driver Score", lambda row: scorecard(row).get("score")),
        ("Coached Events", lambda row: scorecard(row).get("num_coached_events")),
        ("Hard Accels", lambda row: scorecard(row).get("num_hard_accels")),
        ("Hard Brakes", lambda row: scorecard(row).get("num_hard_brakes")),
        ("Hard Corners", lambda row: scorecard(row).get("num_hard_corners")),
        ("Registration Expiry", lambda row: row.get("registration_expiry_date")),
        ("Notes", lambda row: row.get("notes")),
    ]


def _vehicle_columns() -> list[ColumnSpec]:
    return [
        ("Vehicle ID", lambda row: row.get("id")),
        ("Vehicle Number", lambda row: row.get("number")),
        ("Tracked Driver Name", lambda row: _tracked_driver_name(row)),
        ("VIN", lambda row: row.get("vin")),
        ("Make", lambda row: row.get("make")),
        ("Model", lambda row: row.get("model")),
        ("Year", lambda row: row.get("year")),
        ("Status", lambda row: row.get("status")),
        ("Availability Status", lambda row: row.get("availability_status")),
        ("Fuel Type", lambda row: row.get("fuel_type")),
        ("Fuel Level %", lambda row: (row.get("location") or {}).get("fuel_level_percent")),
        ("Fuel Sensor Reading", lambda row: (row.get("location") or {}).get("fuel_sensor_reading")),
        ("Speed MPH", lambda row: (row.get("location") or {}).get("speed_mph")),
        ("Live Address", lambda row: (row.get("location") or {}).get("address")),
        ("City", lambda row: (row.get("location") or {}).get("city")),
        ("State", lambda row: (row.get("location") or {}).get("state")),
        ("Latitude", lambda row: (row.get("location") or {}).get("lat")),
        ("Longitude", lambda row: (row.get("location") or {}).get("lon")),
        ("Located At", lambda row: (row.get("location") or {}).get("located_at")),
        ("Moving Now", lambda row: "Yes" if row.get("is_moving") else "No"),
        ("Stale", lambda row: "Yes" if row.get("is_stale") else "No"),
        ("Active Faults", lambda row: (row.get("fault_summary") or {}).get("active_count")),
        ("Utilization %", lambda row: (row.get("utilization_summary") or {}).get("utilization_percentage")),
        ("Idle Hours 7D", lambda row: _hours((row.get("idle_summary") or {}).get("duration_seconds"))),
        ("Driving Miles 7D", lambda row: (row.get("driving_summary") or {}).get("distance_miles")),
        ("IFTA Miles 30D", lambda row: (row.get("ifta_summary") or {}).get("distance_miles")),
        ("Fuel Purchases 30D", lambda row: (row.get("fuel_purchase_summary") or {}).get("count")),
        ("Inspection Count 30D", lambda row: (row.get("inspection_summary") or {}).get("count")),
        ("Form Count 30D", lambda row: (row.get("form_summary") or {}).get("count")),
        ("ELD Identifier", lambda row: (row.get("eld_device") or {}).get("identifier")),
        ("ELD Model", lambda row: (row.get("eld_device") or {}).get("model")),
    ]


def _fault_columns() -> list[ColumnSpec]:
    return [
        ("Fault ID", lambda row: row.get("id")),
        ("Vehicle Number", lambda row: row.get("vehicle_number")),
        ("Code", lambda row: row.get("code")),
        ("Label", lambda row: row.get("label")),
        ("Description", lambda row: row.get("description")),
        ("Status", lambda row: row.get("status")),
        ("Severity", lambda row: row.get("severity")),
        ("Type", lambda row: row.get("type")),
        ("First Observed", lambda row: row.get("first_observed_at")),
        ("Last Observed", lambda row: row.get("last_observed_at")),
        ("Occurrence Count", lambda row: row.get("occurrence_count")),
        ("Source Address", lambda row: row.get("source_address_label")),
        ("ELD Identifier", lambda row: (row.get("eld_device") or {}).get("identifier")),
        ("ELD Model", lambda row: (row.get("eld_device") or {}).get("model")),
    ]


def _performance_columns() -> list[ColumnSpec]:
    return [
        ("Event ID", lambda row: row.get("id")),
        ("Vehicle Number", lambda row: row.get("vehicle_number")),
        ("Driver Name", lambda row: row.get("driver_name")),
        ("Type", lambda row: row.get("type")),
        ("Primary Behaviors", lambda row: ", ".join(row.get("primary_behaviors") or [])),
        ("Secondary Behaviors", lambda row: ", ".join(row.get("secondary_behaviors") or [])),
        ("Positive Behaviors", lambda row: ", ".join(row.get("positive_behaviors") or [])),
        ("Coaching Status", lambda row: row.get("coaching_status")),
        ("Coached At", lambda row: row.get("coached_at")),
        ("Start Time", lambda row: row.get("start_time")),
        ("End Time", lambda row: row.get("end_time")),
        ("Duration Seconds", lambda row: row.get("duration_seconds")),
        ("Location", lambda row: row.get("location")),
        ("Latitude", lambda row: row.get("lat")),
        ("Longitude", lambda row: row.get("lon")),
        ("Start Speed", lambda row: row.get("start_speed")),
        ("End Speed", lambda row: row.get("end_speed")),
        ("Max Speed", lambda row: row.get("max_speed")),
        ("Min Speed", lambda row: row.get("min_speed")),
        ("Severity", lambda row: row.get("severity")),
        ("Camera Available", lambda row: "Yes" if row.get("camera_available") else "No"),
    ]


def _driving_columns() -> list[ColumnSpec]:
    return [
        ("Trip ID", lambda row: row.get("id")),
        ("Vehicle Number", lambda row: row.get("vehicle_number")),
        ("Driver Name", lambda row: row.get("driver_name")),
        ("Status", lambda row: row.get("status")),
        ("Type", lambda row: row.get("type")),
        ("Start Time", lambda row: row.get("start_time")),
        ("End Time", lambda row: row.get("end_time")),
        ("Duration Seconds", lambda row: row.get("duration_seconds")),
        ("Distance Miles", lambda row: row.get("distance_miles")),
        ("Origin", lambda row: row.get("origin")),
        ("Destination", lambda row: row.get("destination")),
        ("Origin Lat", lambda row: row.get("origin_lat")),
        ("Origin Lon", lambda row: row.get("origin_lon")),
        ("Destination Lat", lambda row: row.get("destination_lat")),
        ("Destination Lon", lambda row: row.get("destination_lon")),
        ("Start Kilometers", lambda row: row.get("start_kilometers")),
        ("End Kilometers", lambda row: row.get("end_kilometers")),
    ]


def _idle_columns() -> list[ColumnSpec]:
    return [
        ("Idle ID", lambda row: row.get("id")),
        ("Vehicle Number", lambda row: row.get("vehicle_number")),
        ("Driver Name", lambda row: row.get("driver_name")),
        ("Start Time", lambda row: row.get("start_time")),
        ("End Time", lambda row: row.get("end_time")),
        ("Duration Seconds", lambda row: row.get("duration_seconds")),
        ("Fuel Start", lambda row: row.get("veh_fuel_start")),
        ("Fuel End", lambda row: row.get("veh_fuel_end")),
        ("Fuel Used", lambda row: row.get("fuel_used")),
        ("City", lambda row: row.get("city")),
        ("State", lambda row: row.get("state")),
        ("Location", lambda row: row.get("location")),
        ("End Type", lambda row: row.get("end_type")),
        ("Latitude", lambda row: row.get("lat")),
        ("Longitude", lambda row: row.get("lon")),
    ]


def _ifta_columns() -> list[ColumnSpec]:
    return [
        ("Trip ID", lambda row: row.get("id")),
        ("Vehicle Number", lambda row: row.get("vehicle_number")),
        ("Date", lambda row: row.get("date")),
        ("Jurisdiction", lambda row: row.get("jurisdiction")),
        ("Distance Miles", lambda row: row.get("distance_miles")),
        ("Start Odometer", lambda row: row.get("start_odometer")),
        ("End Odometer", lambda row: row.get("end_odometer")),
        ("Calibrated Start", lambda row: row.get("calibrated_start_odometer")),
        ("Calibrated End", lambda row: row.get("calibrated_end_odometer")),
        ("Start Lat", lambda row: row.get("start_lat")),
        ("Start Lon", lambda row: row.get("start_lon")),
        ("End Lat", lambda row: row.get("end_lat")),
        ("End Lon", lambda row: row.get("end_lon")),
        ("Time Zone", lambda row: row.get("time_zone")),
    ]


def _fuel_purchase_columns() -> list[ColumnSpec]:
    return [
        ("Purchase ID", lambda row: row.get("id")),
        ("Vehicle Number", lambda row: row.get("vehicle_number")),
        ("Driver Name", lambda row: row.get("driver_name")),
        ("Purchased At", lambda row: row.get("purchased_at")),
        ("Vendor", lambda row: row.get("vendor")),
        ("Location", lambda row: row.get("location")),
        ("Volume", lambda row: row.get("volume")),
        ("Amount", lambda row: row.get("amount")),
        ("Fuel Type", lambda row: row.get("fuel_type")),
    ]


def _inspection_columns() -> list[ColumnSpec]:
    return [
        ("Inspection ID", lambda row: row.get("id")),
        ("Vehicle Number", lambda row: row.get("vehicle_number")),
        ("Driver Name", lambda row: row.get("driver_name")),
        ("Submitted At", lambda row: row.get("submitted_at")),
        ("Safe", lambda row: row.get("safe")),
        ("Inspector", lambda row: row.get("inspector_name")),
        ("Location", lambda row: row.get("location")),
        ("Defects", lambda row: _json_text(row.get("defects"))),
    ]


def _form_columns() -> list[ColumnSpec]:
    return [
        ("Form Entry ID", lambda row: row.get("id")),
        ("Vehicle Number", lambda row: row.get("vehicle_number")),
        ("Driver Name", lambda row: row.get("driver_name")),
        ("Title", lambda row: row.get("title")),
        ("Status", lambda row: row.get("status")),
        ("Submitted At", lambda row: row.get("submitted_at")),
        ("Location", lambda row: row.get("location")),
        ("Answers", lambda row: _json_text(row.get("answers"))),
    ]


def _driver_score_columns() -> list[ColumnSpec]:
    return [
        ("Driver ID", lambda row: row.get("driver_id")),
        ("Driver Name", lambda row: row.get("driver_name")),
        ("Score", lambda row: row.get("score")),
        ("Coached Events", lambda row: row.get("num_coached_events")),
        ("Hard Accels", lambda row: row.get("num_hard_accels")),
        ("Hard Brakes", lambda row: row.get("num_hard_brakes")),
        ("Hard Corners", lambda row: row.get("num_hard_corners")),
        ("Total Kilometers", lambda row: row.get("total_kilometers")),
    ]


def _fleet_status_rank(vehicle: dict, roster_by_name: dict[str, dict]) -> int:
    roster = roster_by_name.get(_normalize_name(_tracked_driver_name(vehicle)), {})
    label = _working_now_label(vehicle, roster)
    ranks = {
        "Moving now": 0,
        "On duty": 1,
        "Active assigned": 2,
        "Parked": 3,
        "Stale": 4,
        "Unassigned": 5,
    }
    return ranks.get(label, 9)


def _movement_label(vehicle: dict) -> str:
    if vehicle.get("is_stale"):
        return "Stale"
    if vehicle.get("is_moving"):
        return "Moving"
    if vehicle.get("location"):
        return "Stopped"
    return "No GPS"


def _working_now_label(vehicle: dict, roster: dict) -> str:
    driver_name = _tracked_driver_name(vehicle)
    duty_status = str(roster.get("duty_status") or "").strip().lower()
    driver_status = str(roster.get("status") or "").strip().lower()

    if vehicle.get("is_moving"):
        return "Moving now"
    if duty_status and duty_status not in {"off_duty", "sleeper_berth", "inactive"}:
        return "On duty"
    if driver_name != "Unassigned" and driver_status == "active" and not vehicle.get("is_stale"):
        return "Active assigned"
    if vehicle.get("is_stale"):
        return "Stale"
    if driver_name == "Unassigned":
        return "Unassigned"
    return "Parked"


def _fuel_state_label(percent: object, sensor: object) -> str:
    if percent not in (None, ""):
        try:
            numeric = float(percent)
        except (TypeError, ValueError):
            numeric = None
        if numeric is not None:
            if numeric <= 10:
                return "Critical fuel"
            if numeric <= 25:
                return "Low fuel"
            if numeric <= 50:
                return "Mid fuel"
            return "Healthy fuel"
    if sensor not in (None, ""):
        return "Sensor only"
    return "Fuel unknown"


def _tracked_driver_name(vehicle: dict) -> str:
    for key in ("driver", "permanent_driver"):
        embedded = vehicle.get(key) or {}
        if embedded.get("full_name"):
            return embedded.get("full_name")
    number = str(vehicle.get("number") or "").strip()
    if "/" in number:
        suffix = number.split("/")[-1].strip(" -")
        if suffix:
            return suffix
    return "Unassigned"


def _tracked_driver_source(vehicle: dict) -> str:
    if (vehicle.get("driver") or {}).get("full_name"):
        return "driver"
    if (vehicle.get("permanent_driver") or {}).get("full_name"):
        return "permanent_driver"
    number = str(vehicle.get("number") or "")
    if "/" in number and number.split("/")[-1].strip(" -"):
        return "vehicle_number_suffix"
    return "unassigned"


def _build_driver_name_map(vehicles: list[dict]) -> dict[str, list[str]]:
    mapping: dict[str, list[str]] = {}
    for vehicle in vehicles:
        driver_name = _tracked_driver_name(vehicle)
        normalized = _normalize_name(driver_name)
        if not normalized or normalized == "unassigned":
            continue
        mapping.setdefault(normalized, []).append(vehicle.get("number") or str(vehicle.get("id") or ""))
    for key in mapping:
        mapping[key] = sorted({value for value in mapping[key] if value})
    return mapping


def _normalize_name(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _estimated_fuel_gallons(percent: object) -> float | None:
    if percent in (None, ""):
        return None
    try:
        numeric = float(percent)
    except (TypeError, ValueError):
        return None
    if numeric < 0:
        numeric = 0
    if numeric > 100:
        numeric = 100
    return round((DEFAULT_TANK_CAPACITY_GALLONS * numeric) / 100, 1)


def _hours(seconds: object) -> float | None:
    if seconds in (None, ""):
        return None
    try:
        return round(float(seconds) / 3600, 2)
    except (TypeError, ValueError):
        return None


def _json_text(value: object) -> str:
    if value in (None, "", [], {}):
        return ""
    return json.dumps(value, ensure_ascii=False)


def _cell_value(value: object):
    if value is None:
        return ""
    if isinstance(value, (int, float, str, bool, datetime)):
        return value
    if isinstance(value, list):
        if not value:
            return ""
        if all(not isinstance(item, (dict, list, tuple, set)) for item in value):
            return ", ".join(str(item) for item in value)
        return _json_text(value)
    if isinstance(value, dict):
        return _json_text(value)
    return str(value)


def _style_header_row(sheet: Worksheet, row_index: int, column_count: int) -> None:
    for column_index in range(1, column_count + 1):
        cell = sheet.cell(row=row_index, column=column_index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGNMENT


def _apply_sheet_formatting(sheet: Worksheet) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = WRAP_ALIGNMENT
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 42)


def _sheet_title(value: str) -> str:
    cleaned = re.sub(r"[\/*?:\[\]]", " ", value).strip()
    return cleaned[:31] or "Sheet"


def _humanize(value: str) -> str:
    return str(value or "").replace("_", " ").strip().title()
